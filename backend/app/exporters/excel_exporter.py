"""Excel 测试用例导出器。

迁移自旧工具 `scripts/generate_excel.py`，面向后端服务调用做了结构化封装。
"""

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from app.ai.models import RequirementItem, TestCaseItem
from app.exporters.adapters import to_legacy_cases, to_legacy_requirements


DEFAULT_COLUMNS = [
    "用例编号",
    "模块名称",
    "用例标题",
    "优先级",
    "关联需求ID",
    "设计方法",
    "前置条件",
    "测试步骤",
    "预期结果",
    "实际结果",
    "是否通过",
    "回归类型",
    "备注",
]

DEFAULT_WIDTHS = [15, 16, 32, 8, 18, 12, 24, 48, 40, 25, 12, 10, 24]

PRIORITY_COLORS = {
    "P0": "FF0000",
    "P1": "FF6600",
    "P2": "FFCC00",
    "P3": "99CC00",
}

FIELD_MAP = {
    "id": "用例编号",
    "编号": "用例编号",
    "case id": "用例编号",
    "tc_id": "用例编号",
    "测试编号": "用例编号",
    "模块": "模块名称",
    "module": "模块名称",
    "功能模块": "模块名称",
    "标题": "用例标题",
    "title": "用例标题",
    "用例名称": "用例标题",
    "测试点": "用例标题",
    "priority": "优先级",
    "级别": "优先级",
    "用例级别": "优先级",
    "req_id": "关联需求ID",
    "需求id": "关联需求ID",
    "requirement": "关联需求ID",
    "需求编号": "关联需求ID",
    "关联需求": "关联需求ID",
    "method": "设计方法",
    "design_method": "设计方法",
    "测试方法": "设计方法",
    "前提条件": "前置条件",
    "precondition": "前置条件",
    "测试前提": "前置条件",
    "步骤": "测试步骤",
    "steps": "测试步骤",
    "操作步骤": "测试步骤",
    "执行步骤": "测试步骤",
    "期望结果": "预期结果",
    "expected": "预期结果",
    "expected result": "预期结果",
    "执行结果": "实际结果",
    "actual": "实际结果",
    "actual result": "实际结果",
    "结果": "是否通过",
    "status": "是否通过",
    "pass/fail": "是否通过",
    "状态": "是否通过",
    "regression": "回归类型",
    "regression_type": "回归类型",
    "回归": "回归类型",
    "说明": "备注",
    "remark": "备注",
    "notes": "备注",
    "其他": "备注",
}


class ExcelExporter:
    def export(
        self,
        cases: list[TestCaseItem],
        output_path: str | Path,
        *,
        requirements: list[RequirementItem] | None = None,
        template_path: str | Path | None = None,
        schema: dict[str, Any] | None = None,
        include_traceability: bool = True,
    ) -> Path:
        output = Path(output_path)
        output.parent.mkdir(parents=True, exist_ok=True)

        legacy_cases = to_legacy_cases(cases)
        legacy_requirements = to_legacy_requirements(requirements or [])
        wb, ws, columns, start_row = self._prepare_workbook(template_path, schema)
        col_index = self._build_column_index(columns)
        self._write_cases(ws, legacy_cases, start_row, col_index)

        end_row = start_row + len(legacy_cases) - 1 if legacy_cases else start_row
        self._add_data_validation(ws, start_row, end_row, col_index)
        self._apply_priority_colors(ws, start_row, end_row, col_index)
        ws.freeze_panes = "A2"

        if include_traceability:
            covered_count, total_req_count = self._create_traceability_sheet(
                wb,
                legacy_cases,
                legacy_requirements,
            )
            self._create_coverage_stats_sheet(wb, legacy_cases, covered_count, total_req_count)

        wb.save(output)
        return output

    def learn_template(self, template_path: str | Path) -> dict[str, Any]:
        template = Path(template_path)
        wb = load_workbook(template)
        ws = wb.active
        columns = []
        widths = []

        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value:
                columns.append(str(cell.value).strip())
                widths.append(ws.column_dimensions[cell.column_letter].width or 12)

        id_format = None
        if ws.max_row >= 2:
            first_id = ws.cell(row=2, column=1).value
            if isinstance(first_id, str):
                if "_" in first_id:
                    id_format = "TC_{MODULE}_{SEQ:03d}"
                elif "-" in first_id:
                    id_format = "{MODULE}-{SEQ:03d}"

        wb.close()
        return {
            "source": str(template),
            "columns": columns,
            "widths": widths,
            "id_format": id_format,
            "learned_at": str(template.stat().st_mtime),
        }

    def _prepare_workbook(
        self,
        template_path: str | Path | None,
        schema: dict[str, Any] | None,
    ):
        if template_path and Path(template_path).exists():
            wb = load_workbook(template_path)
            ws = wb.active
            columns = [cell.value for cell in ws[1] if cell.value]
            start_row = 2
            for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
            return wb, ws, columns, start_row

        columns = schema.get("columns", DEFAULT_COLUMNS) if schema else DEFAULT_COLUMNS
        widths = schema.get("widths", DEFAULT_WIDTHS) if schema else DEFAULT_WIDTHS
        wb = Workbook()
        ws = wb.active
        ws.title = "测试用例"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        border = self._border()

        for col, name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for index, width in enumerate(widths):
            if index < len(columns):
                ws.column_dimensions[chr(65 + index)].width = width

        return wb, ws, columns, 2

    @staticmethod
    def _build_column_index(columns: list[str]) -> dict[str, int]:
        col_index = {}
        for index, col_name in enumerate(columns):
            normalized = str(col_name).lower().strip()
            standard_name = FIELD_MAP.get(normalized, col_name)
            col_index[standard_name] = index + 1
            col_index[col_name] = index + 1
        return col_index

    def _write_cases(self, ws, cases: list[dict], start_row: int, col_index: dict[str, int]):
        border = self._border()
        for row_idx, case in enumerate(cases, start_row):
            for key, value in case.items():
                col_num = col_index.get(key)
                if col_num:
                    cell = ws.cell(row=row_idx, column=col_num, value=value)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    cell.border = border

    @staticmethod
    def _add_data_validation(ws, start_row: int, end_row: int, col_index: dict[str, int]):
        validations = {
            "优先级": '"P0,P1,P2,P3"',
            "回归类型": '"冒烟,核心,全量"',
            "是否通过": '"通过,未通过,阻塞,未执行"',
        }
        for column_name, formula in validations.items():
            if column_name not in col_index:
                continue
            column = col_index[column_name]
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            validation.error = "请选择有效值"
            validation.errorTitle = "无效输入"
            ws.add_data_validation(validation)
            for row in range(start_row, end_row + 1):
                validation.add(ws.cell(row=row, column=column))

    @staticmethod
    def _apply_priority_colors(ws, start_row: int, end_row: int, col_index: dict[str, int]):
        if "优先级" not in col_index:
            return
        priority_col = col_index["优先级"]
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row=row, column=priority_col)
            priority = str(cell.value).upper() if cell.value else ""
            if priority in PRIORITY_COLORS:
                cell.fill = PatternFill(
                    start_color=PRIORITY_COLORS[priority],
                    end_color=PRIORITY_COLORS[priority],
                    fill_type="solid",
                )
                if priority in {"P0", "P1"}:
                    cell.font = Font(bold=True, color="FFFFFF")

    def _create_traceability_sheet(
        self,
        wb,
        cases: list[dict],
        requirements: list[dict],
    ) -> tuple[int, int]:
        ws = wb.create_sheet(title="需求追溯矩阵")
        req_case_map: dict[str, list[str]] = {}
        req_meta = {req["需求ID"]: req for req in requirements if req.get("需求ID")}

        for req_id in req_meta:
            req_case_map.setdefault(req_id, [])

        for case in cases:
            req_ids = [item.strip() for item in str(case.get("关联需求ID") or "").split(",") if item.strip()]
            for req_id in req_ids:
                req_case_map.setdefault(req_id, [])
                case_id = case.get("用例编号")
                if case_id:
                    req_case_map[req_id].append(case_id)

        headers = ["需求ID", "需求名称", "所属模块", "关联用例", "用例数量", "覆盖状态"]
        self._write_header(ws, headers, fill_color="2E7D32")

        covered_count = 0
        for row_idx, (req_id, case_ids) in enumerate(sorted(req_case_map.items()), 2):
            meta = req_meta.get(req_id, {})
            covered = len(case_ids) > 0
            covered_count += 1 if covered else 0
            values = [
                req_id,
                meta.get("需求名称", ""),
                meta.get("所属模块", ""),
                ", ".join(case_ids),
                len(case_ids),
                "已覆盖" if covered else "未覆盖",
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self._border()
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if col_idx == 6 and not covered:
                    cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

        for index, width in enumerate([16, 30, 18, 42, 10, 12], 1):
            ws.column_dimensions[chr(64 + index)].width = width
        ws.freeze_panes = "A2"
        return covered_count, len(req_case_map)

    def _create_coverage_stats_sheet(
        self,
        wb,
        cases: list[dict],
        covered_count: int,
        total_req_count: int,
    ):
        ws = wb.create_sheet(title="覆盖率统计")
        total_cases = len(cases)
        priority_counts = {"P0": 0, "P1": 0, "P2": 0, "P3": 0}
        regression_counts = {"冒烟": 0, "核心": 0, "全量": 0}
        orphan_count = 0

        for case in cases:
            priority = str(case.get("优先级") or "").upper()
            if priority in priority_counts:
                priority_counts[priority] += 1

            regression = case.get("回归类型") or ""
            if regression in regression_counts:
                regression_counts[regression] += 1

            if not case.get("关联需求ID"):
                orphan_count += 1

        coverage_rate = (covered_count / total_req_count * 100) if total_req_count else 0
        coverage_depth = (total_cases / total_req_count) if total_req_count else 0
        stats = [
            ("统计项", "数值"),
            ("总需求数", total_req_count),
            ("已覆盖需求数", covered_count),
            ("未覆盖需求数", total_req_count - covered_count),
            ("需求覆盖率", f"{coverage_rate:.1f}%"),
            ("总用例数", total_cases),
            ("覆盖深度", f"{coverage_depth:.2f} 用例/需求"),
            ("孤儿用例数", orphan_count),
            ("P0 用例数", priority_counts["P0"]),
            ("P1 用例数", priority_counts["P1"]),
            ("P2 用例数", priority_counts["P2"]),
            ("P3 用例数", priority_counts["P3"]),
            ("冒烟测试用例", regression_counts["冒烟"]),
            ("核心回归用例", regression_counts["核心"]),
            ("全量回归用例", regression_counts["全量"]),
        ]

        self._write_header(ws, ["统计项", "数值"], fill_color="1565C0")
        for row_idx, (label, value) in enumerate(stats[1:], 2):
            for col_idx, item in enumerate([label, value], 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=item)
                cell.border = self._border()
                cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center")

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 20

    def _write_header(self, ws, headers: list[str], fill_color: str):
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        for col_idx, name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self._border()

    @staticmethod
    def _border() -> Border:
        return Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
